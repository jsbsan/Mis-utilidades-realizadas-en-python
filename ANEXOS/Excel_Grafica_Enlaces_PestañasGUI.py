import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import re
import networkx as nx
import matplotlib.pyplot as plt
import os

class AnalizadorExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Mapeador de Dependencias Excel")
        self.root.geometry("500x350")
        self.root.resizable(False, False)

        # --- ESTILOS Y COLORES ---
        bg_color = "#f0f0f0"
        self.root.configure(bg=bg_color)
        
        # --- TÍTULO ---
        lbl_titulo = tk.Label(root, text="Analizador de Fórmulas Excel", 
                              font=("Helvetica", 16, "bold"), bg=bg_color, fg="#333")
        lbl_titulo.pack(pady=20)

        # --- INSTRUCCIONES ---
        lbl_info = tk.Label(root, 
                            text="Este programa leerá tu Excel y generará un diagrama\nde las conexiones entre pestañas.",
                            font=("Arial", 10), bg=bg_color, fg="#555")
        lbl_info.pack(pady=10)

        # --- BOTÓN PRINCIPAL ---
        # Usamos un estilo un poco más grande para el botón
        self.btn_cargar = tk.Button(root, text="📂 Seleccionar Archivo Excel", 
                                    font=("Arial", 12, "bold"), 
                                    bg="#4CAF50", fg="white", # Verde
                                    padx=20, pady=10,
                                    cursor="hand2",
                                    command=self.seleccionar_y_procesar)
        self.btn_cargar.pack(pady=20)

        # --- BARRA DE ESTADO / LOG ---
        self.lbl_estado = tk.Label(root, text="Esperando archivo...", 
                                   font=("Consolas", 9), bg="#ddd", relief="sunken", anchor="w")
        self.lbl_estado.pack(side="bottom", fill="x")

    def log(self, mensaje):
        """Actualiza la barra de estado inferior"""
        self.lbl_estado.config(text=f" {mensaje}")
        self.root.update_idletasks() # Forza la actualización visual inmediata

    def seleccionar_y_procesar(self):
        archivo = filedialog.askopenfilename(
            title="Selecciona un archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx *.xlsm")]
        )

        if not archivo:
            return # El usuario canceló

        self.btn_cargar.config(state="disabled") # Desactivar botón mientras procesa
        self.log(f"Leyendo: {os.path.basename(archivo)}...")
        
        # Procesar (llamada a la lógica)
        try:
            dependencias = self.obtener_dependencias(archivo)
            
            if dependencias:
                self.log(f"Encontradas {len(dependencias)} conexiones. Generando archivos...")
                ruta_dot = self.exportar_a_dot(dependencias)
                
                messagebox.showinfo("¡Proceso Terminado!", 
                    f"1. Se ha generado el diagrama visual.\n"
                    f"2. Se guardó el archivo editable:\n{ruta_dot}")
                
                self.graficar_rapido(dependencias)
                self.log("Listo.")
            else:
                messagebox.showinfo("Información", "El archivo no tiene fórmulas que conecten distintas pestañas.")
                self.log("Sin dependencias encontradas.")
                
        except Exception as e:
            messagebox.showerror("Error Crítico", f"Ocurrió un error:\n{str(e)}")
            self.log("Error en el proceso.")
        finally:
            self.btn_cargar.config(state="normal") # Reactivar botón

    def obtener_dependencias(self, file_path):
        wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
        hojas = wb.sheetnames
        conexiones = []
        patron_referencia = re.compile(r"('?)([^'!]+)\1!") # Regex para buscar 'Hoja'!

        for hoja_origen in hojas:
            ws = wb[hoja_origen]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str) and cell.startswith("="):
                        coincidencias = patron_referencia.findall(cell)
                        for match in coincidencias:
                            hoja_destino = match[1]
                            # Evitar auto-referencias y hojas inexistentes
                            if hoja_destino in hojas and hoja_destino != hoja_origen:
                                conexion = (hoja_origen, hoja_destino)
                                if conexion not in conexiones:
                                    conexiones.append(conexion)
        return conexiones

    def exportar_a_dot(self, conexiones):
        filename = "diagrama_excel.dot"
        with open(filename, "w", encoding="utf-8") as f:
            f.write('digraph G {\n')
            f.write('  rankdir="LR";\n')
            f.write('  node [shape=box, style="filled", color="#dbe5f1", fontname="Arial"];\n')
            f.write('  edge [color="#5b9bd5"];\n')
            for origen, destino in conexiones:
                f.write(f'  "{origen}" -> "{destino}";\n')
            f.write('}')
        return os.path.abspath(filename)

    def graficar_rapido(self, conexiones):
        G = nx.DiGraph()
        G.add_edges_from(conexiones)
        plt.figure("Vista Previa del Diagrama", figsize=(10, 8))
        pos = nx.spring_layout(G, k=1.5) # K más alto separa más los nodos
        nx.draw_networkx_nodes(G, pos, node_size=2500, node_color='lightgreen', alpha=0.9)
        nx.draw_networkx_labels(G, pos, font_size=9, font_weight="bold")
        nx.draw_networkx_edges(G, pos, width=2, alpha=0.6, arrowstyle='->', arrowsize=20, connectionstyle="arc3,rad=0.1")
        plt.title("Cierra esta ventana para continuar")
        plt.axis('off')
        plt.show()

if __name__ == "__main__":
    root = tk.Tk()
    app = AnalizadorExcelApp(root)
    root.mainloop()